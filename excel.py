#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
sys.path.append(os.path.dirname(os.path.split(os.path.realpath(__file__))[0]))

import random
import string
import datetime

import xlrd
import xlwt

from openpyxl import Workbook
from openpyxl.style import Alignment

from lib.utils import to_str


class Excel(object):

    def __init__(self, home='/tmp'):
        self.home = home

    def load(self, head, file_name=None, body=None):
        '''
            加载excel
            @head 标题头
            @file_name 文件名
            @body 文件数据
        '''
        if file_name is None and body is None:
            raise Exception("No file name or body")
        # 如果没有文件,则根据body写入文件
        if file_name is None:
            file_name = self.tmp_filename()
            with open(file_name, 'w+b') as codefp:
                codefp.write(body)
        title, data = [], []
        work_book = xlrd.open_workbook(file_name)
        sheet = work_book.sheet_by_index(0)
        for c in xrange(sheet.ncols):
            title.append(sheet.cell(0, c).value)
        for r in xrange(1, sheet.nrows):
            unit = {}
            for c in xrange(sheet.ncols):
                for k, v in head:
                    if title[c].encode('utf-8') == v:
                        cell = sheet.cell(r, c)
                        # 如果是日期类型
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            unit[k] = to_str(self._get_date(cell.value))
                        else:
                            unit[k] = to_str(cell.value)
            data.append(unit)
        os.remove(file_name)
        return data

    def generate(self, title, data, sep=1.5, callback=None):
        '''
            生成EXCEL内容
            @title 标题头
            @data 数据
            @sep 宽度比例
            @callback 回调函数
        '''
        work_book = xlwt.Workbook('UTF-8')
        sheet = work_book.add_sheet('sheet', True)
        ncols = len(title)
        # write title
        title_style = self.get_title_style()
        for j in xrange(ncols):
            sheet.col(j).width = 3333 * sep
            sheet.write(0, j, title[j][1], title_style)
        nrows = len(data)
        text_style = self.get_text_style()
        for i in xrange(nrows):
            for j in xrange(ncols):
                key = title[j][0]
                k_data = data[i]
                # 如果key是对象模式,例如third.book.book_name
                # 则通过取得third对象,再取得book对象,最终得到book_name的值
                while '.' in key:
                    c = key.index(".")
                    k = key[:c]
                    k_data = k_data[k] if key in k_data else getattr(k_data, k)
                    key = key[c + 1:]
                if isinstance(k_data, dict):
                    val = k_data[
                        key] if key in k_data else getattr(k_data, key)
                else:
                    val = k_data[key]
                if isinstance(val, datetime.datetime):
                    val = val.strftime('%Y-%m-%d')
                if callback:
                    val = callback(val)
                sheet.write(i + 1, j, val, text_style)
        filename = self.tmp_filename()
        work_book.save(filename)
        with open(filename, 'rb') as f:
            content = f.read()
        os.remove(filename)
        return content

    def get_title_style(self):
        '''
        excel title style
        '''
        style = xlwt.XFStyle()
        font = xlwt.Font()
        font.bold = True
        style.font = font
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        alignment.vert = xlwt.Alignment.VERT_CENTER
        style.alignment = alignment
        return style

    def get_text_style(self):
        '''
        excel text style
        '''
        style = xlwt.XFStyle()
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        alignment.vert = xlwt.Alignment.VERT_CENTER
        style.alignment = alignment
        return style

    def generate_large(self, title, data):
        '''
        generate excel content
        title: excel title
        data: data result，list
        sep: width ratio
        '''
        wb = Workbook(optimized_write=True)
        # 创建sheet
        ws = wb.create_sheet()
        ws.title = 'sheet'
        # 写头数据
        ws.append([t[1] for t in title])
        ncols, nrows = len(title), len(data)
        for i in xrange(nrows):
            rdata = []
            for j in xrange(ncols):
                key = title[j][0]
                k_data = data[i]
                # 如果key是对象模式,例如third.book.book_name
                # 通过取得third对象,再取得book对象,得到book_name
                while '.' in key:
                    c = key.index(".")
                    k = key[:c]
                    k_data = getattr(k_data, k)
                    key = key[c + 1:]
                if isinstance(k_data, dict):
                    val = k_data[
                        key] if key in k_data else getattr(k_data, key)
                else:
                    val = k_data[key]
                if isinstance(val, datetime.datetime):
                    val = val.strftime('%Y-%m-%d')
                rdata.append(val)
            # 写内容
            ws.append(rdata)
        filename = self.tmp_filename()
        # 保存
        wb.save(filename)
        content = ''
        with open(filename, 'rb') as f:
            content = f.read()
        os.remove(filename)
        return content

    def tmp_filename(self):
        return os.path.join(self.home,
                            ''.join(random.sample(string.lowercase, 10)))

    @staticmethod
    def set_title_style(cel):
        '''
        excel title style
        '''
        cel.style.font.name = 'Arial'
        cel.style.font.bold = True
        cel.style.font.size = 10
        cel.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        cel.style.alignment.vertical = Alignment.VERTICAL_CENTER
        return cel

    @staticmethod
    def set_text_style(cel):
        '''
        excel text style
        '''
        cel.style.font.name = 'Arial'
        cel.style.font.size = 10
        cel.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
        cel.style.alignment.vertical = Alignment.VERTICAL_CENTER
        return cel

    def _get_date(self, vtime):
        '''
            excel日期处理
        '''
        if isinstance(vtime, float):
            vtime = int(vtime)
        s_date = datetime.date(1899, 12, 31).toordinal() - 1
        d = datetime.date.fromordinal(s_date + vtime)
        return d.strftime("%Y-%m-%d")
